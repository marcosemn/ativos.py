import time
from selenium import webdriver
import openpyxl
from openpyxl import load_workbook

chromedriver = r'C:\Users\Dell\Downloads\chromedriver_win32\chromedriver.exe'
 
gc = webdriver.Chrome(executable_path=chromedriver)

wb = load_workbook(r'C:\Users\Dell\Desktop\lista_ativos.xlsx')

def ativo(nome,i):
    print("https://data.anbima.com.br/busca/debentures?q="+nome)
    gc.get("https://data.anbima.com.br/busca/debentures?q="+nome)
    
    time.sleep(5)
    link = gc.find_element_by_xpath("//*[@id='output__container--debentures-item-pu-indicativo-0']/div/span/span/a")
    link.click()
     
    time.sleep(5)
    PUcurva = gc.find_element_by_xpath("//*[@id='precos__table__puhistorico']/tbody/tr[1]/td[3]/span")
    ValorPU = PUcurva.text
    print(ValorPU)
    wb.worksheets[0]["B"+str(i+1)] = ValorPU
    
    evento = gc.find_element_by_xpath("//*[@id='precos__table__puhistorico']/tbody/tr[1]/td[4]/span")
    evento = evento.text
    print(evento)
    wb.worksheets[0]["C"+str(i+1)] = evento
    
    PUmerc = gc.find_element_by_xpath("//*[@id='price-anbima']/table/tbody/tr[1]/td[2]/span")
    PUmerc = PUmerc.text
    print(PUmerc)
    wb.worksheets[0]["D"+str(i+1)] = PUmerc
    
    txmerc = gc.find_element_by_xpath("//*[@id='price-anbima']/table/tbody/tr[1]/td[3]/span")
    txmerc = txmerc.text
    print(txmerc)
    wb.worksheets[0]["E"+str(i+1)] = txmerc

lenght=(wb.worksheets[0].max_row)
print(lenght)

for i in range(1,lenght):
    try:
        nome_ativo=wb.worksheets[0].cell(row=i+1, column=1).value
        ativo(nome_ativo,i)
        time.sleep(5)
    except:
        print('erro')
    
wb.save(r'C:\Users\Dell\Desktop\lista_ativos.xlsx')

    
    
    
    
    
    
    
    